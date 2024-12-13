import pandas as pd
import configparser
import os
from aiapi import process_csv
import db
import openpyxl
from openpyxl.utils import range_boundaries
import csv

config = configparser.ConfigParser()
config.read("config.conf")

def preprocess_xlsx(xlsx_path):
    workbook = openpyxl.load_workbook(xlsx_path)

    if len(workbook.worksheets) != 1:
        # TODO: Add support for multiple sheets
        raise Exception("Only one sheet is allowed in the xlsx file.")
    sheet = workbook.active
    for cell in list(sheet.merged_cells):
        min_col, min_row, max_col, max_row = range_boundaries(str(cell))
        if min_row == max_row:
            continue
        value = sheet.cell(min_row, min_col).value
        sheet.unmerge_cells(str(cell))
        for row in range(min_row, max_row + 1):
            sheet.merge_cells(start_row=row, start_column=min_col, end_row=row, end_column=max_col)
            sheet.cell(row, min_col).value = value
    
    for row in sheet.iter_rows():
        if all(cell.value is None for cell in row):
            sheet.delete_rows(row[0].row)

    for col in sheet.iter_cols():
        if all(cell.value is None for cell in col):
            sheet.delete_cols(col[0].column)

    workbook.save(xlsx_path)

def xlsx_to_csv(xlsx_path):
    data = pd.read_excel(xlsx_path)
    csv_path = config['Directory']['csv_dir'] + "/" + os.path.splitext(os.path.basename(xlsx_path))[0] + ".csv"
    data.to_csv(csv_path, index=False, encoding='utf-8')
    return csv_path

def csv_to_db(csv_path):
    db_path = config['Directory']['db_dir'] + "/" + os.path.splitext(os.path.basename(csv_path))[0] + ".db"
    conn = db.open_db(db_path)
    data = pd.read_csv(csv_path)
    data.to_sql("data", conn, if_exists="replace", index=False)
    conn.close()
    return db_path

def fix_csv(csv_path):
    f = open(csv_path, 'r', encoding='utf-8')
    reader = csv.reader(f)
    rows = list(reader)
    rows = [[cell for cell in row if cell.strip() != ''] for row in rows]
    min_cols = min(len(row) for row in rows)
    rows = [row[:min_cols] for row in rows]
    f.close()
    f = open(csv_path, 'w', encoding='utf-8', newline='')
    writer = csv.writer(f)
    writer.writerows(rows)
    f.close()

def validate(csv_path):
    try:
        f = open(csv_path, 'r', encoding='utf-8')
        reader = csv.reader(f)
        num_cols = len(next(reader))
        for row in reader:
            if len(row) != num_cols:
                raise Exception("Inconsistent number of columns.")
        f.close()
    except Exception as e:
        print("CSV validation failed:", e)
        print("Try to fix the csv file automatically...")
        f.close()
        fix_csv(csv_path)

def xlsx_to_db(xlsx_path):
    print("Starting pre-processing xlsx...")
    preprocess_xlsx(xlsx_path)
    print("finished. Starting xlsx to csv...")
    csv_path = xlsx_to_csv(xlsx_path)
    # csv_path = 'output/csv/0194cc08-946f-4cfd-acee-4baec579fda9.csv'
    print("finished. CSV file:", csv_path)
    print("Starting pre-processing...")
    process_csv(csv_path)
    validate(csv_path)
    print("finished.")
    return csv_to_db(csv_path)

if __name__ == "__main__":
    # preprocess_xlsx('output/xlsx/test.xlsx')
    # xlsx_to_csv('output/xlsx/test.xlsx')
    validate('output/csv/test.csv')
